import openpyxl
from openpyxl.styles import Alignment
import mysql.connector
import sys
from datetime import datetime

# Connection to the database
db = mysql.connector.connect(
    host="localhost",
    user="RP09",
    password="RP09",
    database="rp09"
)

liste_bp = []
cursor = db.cursor(dictionary=True)  # Use of the dictionary cursor for key access

# Retrieving command line arguments
args = sys.argv[1:]

for arg in args:
    cursor.execute("""
            SELECT bonnespratique.num_bp, bonnespratique.test_bp, programme.nom_prog, phase.nom_phase
            FROM appartenance
            JOIN bonnespratique ON appartenance.num_bp = bonnespratique.num_bp
            JOIN programme ON appartenance.num_prog = programme.num_prog
            JOIN phase ON appartenance.num_phase = phase.num_phase
            WHERE bonnespratique.num_bp = %s
        """, (arg,))
    data = cursor.fetchall()
    liste_bp.extend(data)  # Add results to the list

cursor.close()
db.close()

def export_to_excel(liste_bp, creator_name):
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bonnes pratiques"

    # Number of empty columns to add to centre the data
    empty_columns_each_side = 10

    # Starting index for actual data
    data_start_col = empty_columns_each_side + 1

    # Creating the header
    headers = ["ID", "Nom de la bonne pratique", "Programme", "Phase", "Coché"]
    for i, header in enumerate(headers):
        cell = ws.cell(row=1, column=data_start_col + i)
        cell.value = header
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adding data
    for row_index, bp in enumerate(liste_bp, start=2):
        wrapped_test_bp = bp['test_bp']
        row = [bp['num_bp'], wrapped_test_bp, bp['nom_prog'], bp['nom_phase'], " "]
        for col_index, cell_value in enumerate(row):
            cell = ws.cell(row=row_index, column=data_start_col + col_index)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Adjust column widths to display data more clearly
    for i in range(len(headers)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(data_start_col + i)].width = 30

    # Automatically adjust line height to display all text
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        max_height = 0
        for cell in row:
            if cell.value:
                lines = str(cell.value).split('\n')
                max_height = max(max_height, len(lines) * 15)  # Adjust size according to text
        ws.row_dimensions[row[0].row].height = max_height

    # Add text at the bottom of the file
    creation_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    footer_row1 = [""] * 11 + [f"Créé par: {creator_name}"] + [""] * (len(headers) - 1)
    footer_row2 = [""] * 11 + [f"Date de création: {creation_date}"] + [""] * (len(headers) - 1)
    ws.append([""] * (empty_columns_each_side + len(headers)))  # Empty line
    ws.append(footer_row1)
    ws.append(footer_row2)

    # Centre the content of footer cells
    for row in ws.iter_rows(min_row=ws.max_row - 1, max_row=ws.max_row, min_col=data_start_col, max_col=data_start_col + len(headers) - 1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save Excel file
    wb.save("bonnes_pratiques.xlsx")
    print("Excel généré avec succès !")

# Name of creator (replace with real name)
creator_name = "Mathis"

export_to_excel(liste_bp, creator_name)









